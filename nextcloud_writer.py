"""
Nextcloud daily file writer for Alleyne Group bid scanner.
Writes one Excel file per platform per profile per day.
Naming: DATE_PLATFORM_Company_ProfileName.xlsx

Master BidTracker is handled by combiner.py
"""

import os
import io
import re
import logging
from datetime import datetime

import requests
from requests.auth import HTTPBasicAuth
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

NEXTCLOUD_URL      = os.getenv("NEXTCLOUD_URL",    "https://cloud.alleyneinc.net")
NEXTCLOUD_USER     = os.getenv("NEXTCLOUD_USER",   "tzvorygina")
NEXTCLOUD_PASSWORD = os.getenv("NEXTCLOUD_PASSWORD", "")
NEXTCLOUD_FOLDER   = os.getenv("NEXTCLOUD_FOLDER", "Alleyne Inc/AlleyneAdmAgent")

COLOR_URGENT   = "FFCCCC"
COLOR_SOON     = "FFF3CC"
COLOR_CLIENT   = "CCE5FF"
COLOR_HOT      = "E8CCFF"
COLOR_POSSIBLE = "CCEECC"
COLOR_WEAK     = "EEEEEE"
COLOR_HEADER   = "2D3748"

COLUMNS = [
    ("Source",               12),
    ("Opportunity",          42),
    ("Client / Account",     30),
    ("Solicitation #",       18),
    ("Reference #",          18),
    ("Profile",              20),
    ("Solicitation Type",    20),
    ("AI Amount Guess",      18),
    ("Closing Date",         20),
    ("Days Left",            10),
    ("Published Date",       20),
    ("Location",             25),
    ("Contract Duration",    18),
    ("Bid Intent",           12),
    ("Quick Summary",        50),
    ("Contact Name",         20),
    ("Contact Email",        25),
    ("RFP URL",              35),
    ("Agreement Types",      25),
    ("Date Found",           18),
    ("Platform ID",          20),
    ("Relevance Score",      15),
    ("Matched Capabilities", 35),
    ("Matched Signals",      30),
]

def _auth():
    return HTTPBasicAuth(NEXTCLOUD_USER, NEXTCLOUD_PASSWORD)

def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def _border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def make_filename(company, profile_name, platform):
    today = datetime.now().strftime("%Y-%m-%d")
    company_clean = re.sub(r'[^\w]', '', company)[:15]
    profile_clean = re.sub(r'[^\w-]', '-', profile_name)[:25].strip('-')
    return f"{today}_{platform}_{company_clean}_{profile_clean}.xlsx"

def row_color(opp, known_clients):
    days  = opp.get("days_to_close", 999)
    score = opp.get("score", 0)
    org   = (opp.get("organization", "") + " " + opp.get("issuing_org", "")).lower()
    is_known = any(
        any(w in org for w in c["name"].lower().split() if len(w) > 4)
        for c in known_clients
    )
    if days <= 3: return COLOR_URGENT
    if days <= 7: return COLOR_SOON
    if is_known:  return COLOR_CLIENT
    if score >= 60: return COLOR_HOT
    if score >= 35: return COLOR_POSSIBLE
    return COLOR_WEAK

def guess_amount(opp):
    duration = (opp.get("contract_duration", "") or "").lower()
    sol_type = (opp.get("solicitation_type", "") or "").lower()
    org      = (opp.get("organization", "") or "").lower()
    title    = (opp.get("title", "") or "").lower()
    years = 1
    m = re.search(r'(\d+)\s*year', duration)
    if m: years = int(m.group(1))
    if "rfsa" in sol_type or "supply arrangement" in sol_type: base = 500_000
    elif "rfp" in sol_type and "formal" in sol_type: base = 200_000
    elif "rfp" in sol_type: base = 150_000
    elif "rfq" in sol_type: base = 100_000
    else: base = 75_000
    mult = 1.0
    if any(k in org for k in ["federal", "canada", "department"]): mult = 1.5
    elif any(k in org for k in ["university", "hospital", "bank", "insurance"]): mult = 1.3
    if any(k in title for k in ["enterprise", "transformation", "erp"]): mult *= 1.4
    return f"~${max(round(base * years * mult / 25_000) * 25_000, 75_000):,.0f}"

def build_workbook(opportunities, known_clients, company, profile_name, platform, scan_date):
    wb = openpyxl.Workbook()
    ws = wb.active
    # Excel sheet names cannot contain: / \ ? * [ ] :
    safe_title = f"{platform}-{profile_name}"
    for ch in "/\\?*[]:":
        safe_title = safe_title.replace(ch, "-")
    ws.title = safe_title[:31]

    total  = len(opportunities)
    hot    = sum(1 for o in opportunities if o.get("score", 0) >= 60)
    urgent = sum(1 for o in opportunities if o.get("days_to_close", 999) <= 3)

    ws.merge_cells(f"A1:{get_column_letter(len(COLUMNS))}1")
    c = ws.cell(row=1, column=1,
                value=f"{platform} | {company} | {profile_name} | {scan_date} | {total} opps | HOT: {hot} | URGENT: {urgent}")
    c.font = Font(bold=True, color="FFFFFF", size=12)
    c.fill = _fill(COLOR_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    stats = [f"Total: {total}", f"HOT(60+): {hot}", f"Possible(35-59): {sum(1 for o in opportunities if 35<=o.get('score',0)<60)}",
             f"Urgent(≤3d): {urgent}", "💜 Purple=HOT  🌿 Green=Possible  🌸 Rose=Urgent  🌼 Yellow=Soon"]
    for i, s in enumerate(stats, 1):
        ws.cell(row=2, column=i, value=s).font = Font(size=9, italic=True)
    ws.row_dimensions[2].height = 14

    for col_idx, (name, width) in enumerate(COLUMNS, 1):
        c = ws.cell(row=3, column=col_idx, value=name)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = _fill(COLOR_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[3].height = 28
    ws.freeze_panes = "C4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(COLUMNS))}3"

    sorted_opps = sorted(opportunities, key=lambda x: x.get("score", 0), reverse=True)

    for row_idx, opp in enumerate(sorted_opps, 4):
        fill = _fill(row_color(opp, known_clients))
        sources = opp.get("sources", [platform])
        source_label = " + ".join(sorted(set(sources)))
        pid = (opp.get("platform_id", "")
               or f"{sources[0] if sources else platform}:{opp.get('merx_id') or opp.get('biddingo_id') or opp.get('solicitation_number','')}")

        vals = [
            source_label,
            opp.get("title", ""),
            opp.get("organization", ""),
            opp.get("solicitation_number", ""),
            opp.get("reference_number", ""),
            opp.get("profile", profile_name),
            opp.get("solicitation_type", ""),
            guess_amount(opp),
            opp.get("closing_date", ""),
            opp.get("days_to_close", ""),
            opp.get("published_date", ""),
            opp.get("location", ""),
            opp.get("contract_duration", ""),
            opp.get("bid_intent", ""),
            (opp.get("description") or "")[:300],
            opp.get("contact_name", ""),
            opp.get("contact_email", ""),
            opp.get("url", ""),
            ", ".join(opp.get("agreement_types") or []),
            scan_date[:10],
            pid,
            opp.get("score", 0),
            ", ".join((opp.get("matched_capabilities") or [])[:3]),
            ", ".join((opp.get("matched_signals") or [])[:3]),
        ]

        for col_idx, val in enumerate(vals, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.fill = fill
            c.border = _border()
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.font = Font(size=9)
        ws.row_dimensions[row_idx].height = 40

    return wb

def upload_to_nextcloud(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    folder_url = f"{NEXTCLOUD_URL}/remote.php/dav/files/{NEXTCLOUD_USER}/{NEXTCLOUD_FOLDER}"
    requests.request("MKCOL", folder_url, auth=_auth())
    url = f"{NEXTCLOUD_URL}/remote.php/dav/files/{NEXTCLOUD_USER}/{NEXTCLOUD_FOLDER}/{filename}"
    r = requests.put(url, data=buf.getvalue(), auth=_auth(),
                     headers={"Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"})
    if r.status_code not in (200, 201, 204):
        raise Exception(f"Upload failed: {r.status_code} {r.text}")
    log.info(f"Uploaded: {filename}")
    return url

def write_to_nextcloud(opportunities, known_clients,
                       company="Alleyne Inc.",
                       profile_name="All-Profiles",
                       platform="MerxS") -> str:
    scan_date = datetime.now().strftime("%Y-%m-%d %H:%M")
    log.info(f"Writing daily file: {platform}/{company}/{profile_name} — {len(opportunities)} opportunities")
    wb = build_workbook(opportunities, known_clients, company, profile_name, platform, scan_date)
    filename = make_filename(company, profile_name, platform)
    url = upload_to_nextcloud(wb, filename)
    return url
