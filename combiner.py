"""
Combiner Agent — Alleyne Group
Reads all daily scan files from Nextcloud, merges, deduplicates,
updates master BidTracker.xlsx on Nextcloud, then mirrors to Google Drive
so Claude can read results each session.
Runs after all platform scanners complete.

FIXES 2026-04-29:
  - Bug 4: read_daily_file() now joins all filename parts from index 3 onward
            to get full profile name (e.g. "IT-IM-KM-Profile" not just "IT")
"""

import os
import io
import re
import json
import base64
import logging
import requests
from datetime import datetime, date
from typing import Optional

from requests.auth import HTTPBasicAuth
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

# ── Config ────────────────────────────────────────────────────────────────────
NEXTCLOUD_URL      = os.getenv("NEXTCLOUD_URL",    "https://cloud.alleyneinc.net")
NEXTCLOUD_USER     = os.getenv("NEXTCLOUD_USER",   "tzvorygina")
NEXTCLOUD_PASSWORD = os.getenv("NEXTCLOUD_PASSWORD", "")
NEXTCLOUD_FOLDER   = os.getenv("NEXTCLOUD_FOLDER", "Alleyne Inc/AlleyneAdmAgent")
MASTER_FILENAME    = "AlleyneInc_BidTracker.xlsx"
KNOWN_CLIENTS_FILE = os.getenv("KNOWN_CLIENTS", "/app/known_clients.json")

# Google Drive folder: "BidScanner Results" inside "Alleyne Inc AI Agents - Execution"
GDRIVE_FOLDER_ID   = os.getenv("GDRIVE_FOLDER_ID", "1z140EEcsGvOvr2w-UHuByvL9h33lmNL7")
GDRIVE_TOKEN       = os.getenv("GDRIVE_TOKEN", "")  # OAuth token if available

# ── Colors ────────────────────────────────────────────────────────────────────
COLOR_URGENT   = "FFCCCC"
COLOR_SOON     = "FFF3CC"
COLOR_CLIENT   = "CCE5FF"
COLOR_HOT      = "E8CCFF"
COLOR_POSSIBLE = "CCEECC"
COLOR_WEAK     = "EEEEEE"
COLOR_EXPIRED  = "D0D0D0"
COLOR_HEADER   = "2D3748"

# ── Column definitions ────────────────────────────────────────────────────────
AGENT_COLUMNS = [
    ("Source",               12),
    ("Flags",                18),
    ("Status",               14),
    ("Client / Account",     30),
    ("Opportunity",          42),
    ("Reference #",          18),
    ("Solicitation #",       18),
    ("Profile",              20),
    ("Solicitation Type",    20),
    ("AI Amount Guess",      18),
    ("Closing Date",         20),
    ("Days Left",            10),
    ("Published Date",       20),
    ("Location",             25),
    ("Contract Duration",    18),
    ("Bid Intent",           12),
    ("Quick Summary",        50),
    ("Contact Name",         20),
    ("Contact Email",        25),
    ("RFP URL",              35),
    ("Agreement Types",      25),
    ("Date Found",           18),
    ("Platform ID",          20),
    ("Relevance Score",      15),
    ("Matched Capabilities", 35),
    ("Matched Signals",      30),
]

HUMAN_COLUMNS = [
    ("Sales Stage",          15),
    ("Amount (Est.)",        15),
    ("Probability %",        13),
    ("Weighted Value",       15),
    ("Bid Decision",         15),
    ("Notes",                40),
    ("Assigned To",          18),
    ("RFP Questions",        30),
    ("Reasons for Passing",  30),
]

ALL_COLUMNS = AGENT_COLUMNS + HUMAN_COLUMNS

# ── Helpers ───────────────────────────────────────────────────────────────────
def _auth():
    return HTTPBasicAuth(NEXTCLOUD_USER, NEXTCLOUD_PASSWORD)

def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def _border() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def load_known_clients() -> list:
    try:
        with open(KNOWN_CLIENTS_FILE) as f:
            return json.load(f).get("known_clients", [])
    except Exception:
        return []

def guess_amount(opp: dict) -> str:
    duration = (opp.get("contract_duration", "") or "").lower()
    sol_type = (opp.get("solicitation_type", "") or "").lower()
    org      = (opp.get("organization", "") or "").lower()
    title    = (opp.get("title", "") or "").lower()
    years = 1
    m = re.search(r'(\d+)\s*year', duration)
    if m:
        years = int(m.group(1))
    if "rfsa" in sol_type or "supply arrangement" in sol_type:
        base = 500_000
    elif "rfp" in sol_type and "formal" in sol_type:
        base = 200_000
    elif "rfp" in sol_type:
        base = 150_000
    elif "rfq" in sol_type:
        base = 100_000
    else:
        base = 75_000
    mult = 1.0
    if any(k in org for k in ["federal", "canada", "department", "government of canada"]):
        mult = 1.5
    elif any(k in org for k in ["university", "hospital", "bank", "insurance"]):
        mult = 1.3
    if any(k in title for k in ["enterprise", "transformation", "erp", "platform"]):
        mult *= 1.4
    estimate = round(base * years * mult / 25_000) * 25_000
    return f"~${max(estimate, 75_000):,.0f}"

def row_color(opp: dict, known_clients: list, status: str) -> str:
    if status in ("EXPIRED", "REJECTED"):
        return COLOR_EXPIRED
    days  = opp.get("days_to_close", 999)
    score = opp.get("score", 0)
    org   = (opp.get("organization", "") + " " + opp.get("issuing_org", "")).lower()
    is_known = any(
        any(w in org for w in c["name"].lower().split() if len(w) > 4)
        for c in known_clients
    )
    if days <= 3:   return COLOR_URGENT
    if days <= 7:   return COLOR_SOON
    if is_known:    return COLOR_CLIENT
    if score >= 60: return COLOR_HOT
    if score >= 35: return COLOR_POSSIBLE
    return COLOR_WEAK

def build_flags(opp: dict, known_clients: list) -> str:
    flags   = []
    days    = opp.get("days_to_close", 999)
    score   = opp.get("score", 0)
    sources = opp.get("sources", [])
    if days <= 3:   flags.append("URGENT")
    elif days <= 7: flags.append("SOON")
    org = (opp.get("organization", "") + " " + opp.get("issuing_org", "")).lower()
    for client in known_clients:
        name = client["name"].lower()
        if any(w in org for w in name.split() if len(w) > 4):
            flags.append("CLIENT★" if client.get("won") else "CLIENT")
            break
    if score >= 60:      flags.append("HOT")
    if len(sources) > 1: flags.append("MULTI")
    return " + ".join(flags) if flags else ""

# ── Nextcloud file operations ─────────────────────────────────────────────────
def list_today_files() -> list:
    today       = date.today().strftime("%Y-%m-%d")
    webdav_base = f"{NEXTCLOUD_URL}/remote.php/dav/files/{NEXTCLOUD_USER}/{NEXTCLOUD_FOLDER}"
    r = requests.request("PROPFIND", webdav_base, auth=_auth(),
                         headers={"Depth": "1", "Content-Type": "application/xml"})
    if r.status_code not in (207, 200):
        log.warning(f"Could not list Nextcloud folder: {r.status_code} {r.text[:200]}")
        return []
    log.info(f"PROPFIND response snippet: {r.text[:500]}")
    files = []
    for pattern in [r'<D:href>([^<]+)</D:href>', r'<d:href>([^<]+)</d:href>', r'<href>([^<]+)</href>']:
        for match in re.finditer(pattern, r.text, re.IGNORECASE):
            path = match.group(1)
            try:
                from urllib.parse import unquote
                path = unquote(path)
            except Exception:
                pass
            filename = path.split("/")[-1]
            if filename.startswith(today) and filename.endswith(".xlsx") and filename != MASTER_FILENAME:
                if filename not in files:
                    files.append(filename)
    log.info(f"Found {len(files)} daily files for {today}: {files}")
    return files

def download_file(filename: str) -> Optional[openpyxl.Workbook]:
    url = f"{NEXTCLOUD_URL}/remote.php/dav/files/{NEXTCLOUD_USER}/{NEXTCLOUD_FOLDER}/{filename}"
    r   = requests.get(url, auth=_auth())
    if r.status_code == 200:
        return openpyxl.load_workbook(io.BytesIO(r.content))
    log.warning(f"Could not download {filename}: {r.status_code}")
    return None

def download_master() -> Optional[openpyxl.Workbook]:
    url = f"{NEXTCLOUD_URL}/remote.php/dav/files/{NEXTCLOUD_USER}/{NEXTCLOUD_FOLDER}/{MASTER_FILENAME}"
    r   = requests.get(url, auth=_auth())
    if r.status_code == 200:
        log.info("Downloaded existing master file")
        return openpyxl.load_workbook(io.BytesIO(r.content))
    log.info("Master file not found — will create new")
    return None

def upload_master(wb: openpyxl.Workbook) -> str:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    folder_url = f"{NEXTCLOUD_URL}/remote.php/dav/files/{NEXTCLOUD_USER}/{NEXTCLOUD_FOLDER}"
    requests.request("MKCOL", folder_url, auth=_auth())
    url = f"{NEXTCLOUD_URL}/remote.php/dav/files/{NEXTCLOUD_USER}/{NEXTCLOUD_FOLDER}/{MASTER_FILENAME}"
    r   = requests.put(url, data=buf.getvalue(), auth=_auth(),
                       headers={"Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"})
    if r.status_code not in (200, 201, 204):
        raise Exception(f"Upload failed: {r.status_code} {r.text}")
    log.info(f"Master uploaded: {url}")
    return url

# ── Google Drive upload ───────────────────────────────────────────────────────
def upload_to_gdrive(wb: openpyxl.Workbook, scan_date: str):
    """
    Upload master BidTracker to Google Drive BidScanner Results folder.
    Falls back to logging a warning if no token available.
    """
    token = GDRIVE_TOKEN
    if not token:
        log.warning("GDRIVE_TOKEN not set — skipping Google Drive upload. "
                    "Set GDRIVE_TOKEN in .env to enable Drive mirroring.")
        return

    buf = io.BytesIO()
    wb.save(buf)
    file_bytes  = buf.getvalue()
    today       = date.today().strftime("%Y-%m-%d")
    gdrive_name = f"{today}_AlleyneInc_BidTracker.xlsx"

    search_url = "https://www.googleapis.com/drive/v3/files"
    headers    = {"Authorization": f"Bearer {token}"}
    search_r   = requests.get(search_url, headers=headers, params={
        "q": f"name='{gdrive_name}' and '{GDRIVE_FOLDER_ID}' in parents and trashed=false",
        "fields": "files(id,name)"
    })
    if search_r.status_code == 200:
        existing = search_r.json().get("files", [])
        for f in existing:
            requests.delete(f"{search_url}/{f['id']}", headers=headers)
            log.info(f"Deleted old Drive file: {f['name']}")

    metadata = json.dumps({
        "name":    gdrive_name,
        "parents": [GDRIVE_FOLDER_ID],
    })
    upload_r = requests.post(
        "https://www.googleapis.com/upload/drive/v3/files?uploadType=multipart",
        headers={"Authorization": f"Bearer {token}"},
        files={
            "metadata": ("metadata", metadata, "application/json; charset=UTF-8"),
            "file":     ("file", file_bytes,
                         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        }
    )
    if upload_r.status_code in (200, 201):
        file_id = upload_r.json().get("id", "")
        log.info(f"Uploaded to Google Drive: {gdrive_name} (id: {file_id})")
    else:
        log.warning(f"Google Drive upload failed: {upload_r.status_code} {upload_r.text[:200]}")

# ── Read opportunities from daily xlsx files ──────────────────────────────────
def read_daily_file(wb: openpyxl.Workbook, filename: str) -> list:
    """
    Extract opportunity dicts from a daily scan file.

    Filename format: 2026-04-28_BidDS_AlleyneInc_IT-IM-KM-Profile.xlsx
      parts[0] = date
      parts[1] = platform code (BidDS, MerxS, BonfireS)
      parts[2] = company
      parts[3:] = profile name parts — FIX Bug 4: join all to get full profile name

    Row structure:
      Row 1: title banner
      Row 2: stats line
      Row 3: column headers (detected dynamically)
      Row 4+: data
    """
    parts    = filename.replace(".xlsx", "").split("_")
    platform = parts[1] if len(parts) > 1 else "Unknown"
    # FIX Bug 4: join all parts from index 3 to get full profile name
    # e.g. "IT-IM-KM-Profile" instead of just "IT"
    profile  = "_".join(parts[3:]) if len(parts) > 3 else "Unknown"

    ws      = wb.active
    opps    = []
    headers = []
    header_row_found = False

    for row in ws.iter_rows(values_only=True):
        if not header_row_found:
            row_vals = [str(c).strip() if c else "" for c in row]
            if "Opportunity" in row_vals or "Source" in row_vals:
                headers = row_vals
                header_row_found = True
            continue

        if not any(row):
            continue

        opp = {}
        for col_idx, val in enumerate(row):
            if col_idx < len(headers) and headers[col_idx]:
                opp[headers[col_idx]] = val

        mapped = {
            "title"               : opp.get("Opportunity") or opp.get("Title") or "",
            "organization"        : opp.get("Client / Account") or opp.get("Organization") or "",
            "issuing_org"         : opp.get("Client / Account") or "",
            "solicitation_number" : opp.get("Solicitation #") or opp.get("Solicitation Number") or "",
            "reference_number"    : opp.get("Reference #") or "",
            "solicitation_type"   : opp.get("Solicitation Type") or "",
            "closing_date"        : str(opp.get("Closing Date") or ""),
            "days_to_close"       : int(opp.get("Days Left") or 0),
            "published_date"      : str(opp.get("Published Date") or ""),
            "location"            : opp.get("Location") or "",
            "description"         : opp.get("Quick Summary") or "",
            "contact_name"        : opp.get("Contact Name") or "",
            "contact_email"       : opp.get("Contact Email") or "",
            "url"                 : opp.get("RFP URL") or "",
            "agreement_types"     : [],
            "contract_duration"   : opp.get("Contract Duration") or "",
            "bid_intent"          : opp.get("Bid Intent") or "",
            "bid_submission_type" : "",
            "qa_deadline"         : "",
            "score"               : int(opp.get("Relevance Score") or 0),
            "matched_capabilities": [c.strip() for c in str(opp.get("Matched Capabilities") or "").split(",") if c.strip()],
            "matched_signals"     : [s.strip() for s in str(opp.get("Matched Signals") or "").split(",") if s.strip()],
            "sources"             : [opp.get("Source") or platform],
            "profile"             : opp.get("Profile") or profile,
            "platform_id"         : opp.get("Platform ID") or "",
            "recommendation"      : "",
        }

        if not mapped["title"]:
            continue
        opps.append(mapped)

    log.info(f"Read {len(opps)} opportunities from {filename} (platform={platform}, profile={profile})")
    return opps

# ── Merge opportunities ───────────────────────────────────────────────────────
def merge_opportunities(all_opps: list) -> list:
    merged = {}
    for opp in all_opps:
        sol_num = (opp.get("solicitation_number") or "").strip().upper()
        title   = (opp.get("title") or "").strip().lower()[:60]
        key     = sol_num if len(sol_num) > 3 else title
        if not key:
            merged[id(opp)] = opp
            continue
        if key in merged:
            existing = merged[key]
            combined = list(set(existing.get("sources", []) + opp.get("sources", [])))
            existing["sources"] = combined
            if len(combined) > 1:
                existing["score"] = min(100, existing.get("score", 0) + 5)
                log.info(f"Merged cross-platform: {key} {combined}")
        else:
            merged[key] = opp
    return list(merged.values())

# ── Read existing master rows ─────────────────────────────────────────────────
def read_master_rows(ws) -> dict:
    rows       = {}
    pid_col    = next((i+1 for i, (n,_) in enumerate(AGENT_COLUMNS) if n == "Platform ID"), None)
    status_col = next((i+1 for i, (n,_) in enumerate(AGENT_COLUMNS) if n == "Status"), None)
    if not pid_col or not status_col:
        return rows
    for row in ws.iter_rows(min_row=5):
        pid_cell    = row[pid_col - 1] if pid_col <= len(row) else None
        status_cell = row[status_col - 1] if status_col <= len(row) else None
        if not pid_cell or not pid_cell.value:
            continue
        pid         = str(pid_cell.value)
        status      = str(status_cell.value) if status_cell and status_cell.value else "ACTIVE"
        human_start = len(AGENT_COLUMNS)
        human_vals  = [row[human_start + i].value if (human_start + i) < len(row) else ""
                       for i in range(len(HUMAN_COLUMNS))]
        rows[pid]   = {"row": pid_cell.row, "status": status, "human_vals": human_vals}
    return rows

# ── Write master workbook ─────────────────────────────────────────────────────
def write_master_sheet(ws, opportunities: list, known_clients: list,
                       existing_rows: dict, scan_date: str, active_platforms: set):
    ws.delete_rows(1, ws.max_row)
    total  = len(opportunities)
    active = sum(1 for o in opportunities if o.get("_status") not in ("EXPIRED", "REJECTED"))

    ws.merge_cells(f"A1:{get_column_letter(len(ALL_COLUMNS))}1")
    c = ws.cell(row=1, column=1,
                value=f"ALLEYNE GROUP — BID TRACKER | Last scan: {scan_date} | Active: {active} | Total: {total}")
    c.font      = Font(bold=True, color="FFFFFF", size=13)
    c.fill      = _fill(COLOR_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 6

    legend = ["🌸 Rose=Urgent(≤3d)", "🌼 Yellow=Soon(≤7d)", "💙 Blue=Known client",
              "💜 Purple=HOT(60+)", "🌿 Green=Possible(35+)", "⬜ Gray=Weak(<35)"]
    for i, txt in enumerate(legend, 1):
        c = ws.cell(row=3, column=i, value=txt)
        c.font = Font(size=9, italic=True)

    for col_idx, (name, width) in enumerate(ALL_COLUMNS, 1):
        c = ws.cell(row=4, column=col_idx, value=name)
        c.font      = Font(bold=True, color="FFFFFF", size=10)
        c.fill      = _fill(COLOR_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[4].height = 30
    ws.freeze_panes = "C5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(ALL_COLUMNS))}4"

    def sort_key(o):
        s     = o.get("_status", "ACTIVE")
        order = {"NEW": 0, "ACTIVE": 1, "EXPIRED": 9, "REJECTED": 9}.get(s, 5)
        return (order, -o.get("score", 0))

    data_row = 5
    for opp in sorted(opportunities, key=sort_key):
        status       = opp.get("_status", "ACTIVE")
        flags        = build_flags(opp, known_clients)
        fill         = _fill(row_color(opp, known_clients, status))
        source_label = " + ".join(sorted(set(opp.get("sources", []))))
        pid          = opp.get("platform_id", "")

        agent_vals = [
            source_label, flags, status,
            opp.get("organization", ""), opp.get("title", ""),
            opp.get("reference_number", ""), opp.get("solicitation_number", ""),
            opp.get("profile", ""), opp.get("solicitation_type", ""),
            guess_amount(opp), opp.get("closing_date", ""),
            opp.get("days_to_close", ""), opp.get("published_date", ""),
            opp.get("location", ""), opp.get("contract_duration", ""),
            opp.get("bid_intent", ""), (opp.get("description") or "")[:300],
            opp.get("contact_name", ""), opp.get("contact_email", ""),
            opp.get("url", ""),
            ", ".join(opp.get("agreement_types") or []),
            scan_date[:10], pid, opp.get("score", 0),
            ", ".join((opp.get("matched_capabilities") or [])[:3]),
            ", ".join((opp.get("matched_signals") or [])[:3]),
        ]

        human_vals = existing_rows.get(pid, {}).get("human_vals", [""] * len(HUMAN_COLUMNS))
        for col_idx, val in enumerate(agent_vals + human_vals, 1):
            c           = ws.cell(row=data_row, column=col_idx, value=val)
            c.fill      = fill
            c.border    = _border()
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.font      = Font(size=9, color="888888" if status in ("EXPIRED", "REJECTED") else "000000")
        ws.row_dimensions[data_row].height = 45
        data_row += 1

# ── Main combiner function ────────────────────────────────────────────────────
def run_combiner(active_platforms: set = None) -> str:
    log.info("=== Running Combiner ===")
    scan_date     = datetime.now().strftime("%Y-%m-%d %H:%M")
    known_clients = load_known_clients()

    # 1. Find today's daily files
    daily_files = list_today_files()
    if not daily_files:
        log.warning("No daily files found for today — combiner has nothing to merge")
        return ""

    # 2. Read all opportunities
    all_opps = []
    for filename in daily_files:
        wb = download_file(filename)
        if wb:
            all_opps.extend(read_daily_file(wb, filename))
    log.info(f"Total opportunities from daily files: {len(all_opps)}")

    # 3. Merge cross-platform duplicates
    merged = merge_opportunities(all_opps)
    log.info(f"After merge: {len(all_opps)} → {len(merged)}")

    # 4. Download existing master
    master_wb = download_master()
    if master_wb is None:
        master_wb             = openpyxl.Workbook()
        master_wb.active.title = "Active Bids"

    if "Active Bids" not in master_wb.sheetnames:
        master_wb.create_sheet("Active Bids", 0)
    if "Archive" not in master_wb.sheetnames:
        master_wb.create_sheet("Archive")

    ws_active  = master_wb["Active Bids"]
    ws_archive = master_wb["Archive"]

    # 5. Read existing rows to preserve human columns
    existing_rows = read_master_rows(ws_active)
    incoming_pids = {opp.get("platform_id", "") for opp in merged if opp.get("platform_id")}

    # 6. Determine status
    for opp in merged:
        pid = opp.get("platform_id", "")
        if pid in existing_rows:
            opp["_status"] = "REJECTED" if existing_rows[pid]["status"] == "REJECTED" else "ACTIVE"
        else:
            opp["_status"] = "NEW"

    # Mark expired
    if active_platforms:
        for pid, info in existing_rows.items():
            if info["status"] in ("EXPIRED", "REJECTED"):
                continue
            row_platform = pid.split(":")[0] if ":" in pid else ""
            if row_platform in active_platforms and pid not in incoming_pids:
                log.info(f"Marking expired: {pid}")
                merged.append({
                    "title": pid, "organization": "", "score": 0,
                    "days_to_close": 0, "sources": [row_platform],
                    "platform_id": pid, "_status": "EXPIRED",
                    "profile": "", "solicitation_number": "",
                })

    # 7. Archive sheet headers
    if ws_archive.max_row <= 1:
        for col_idx, (name, _) in enumerate(ALL_COLUMNS, 1):
            c      = ws_archive.cell(row=1, column=col_idx, value=name)
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.fill = _fill(COLOR_HEADER)
        ws_archive.freeze_panes = "C2"

    # 8. Split and write
    active_opps  = [o for o in merged if o.get("_status") != "EXPIRED"]
    expired_opps = [o for o in merged if o.get("_status") == "EXPIRED"]

    write_master_sheet(ws_active, active_opps, known_clients,
                       existing_rows, scan_date, active_platforms or set())

    arch_row = ws_archive.max_row + 1
    for opp in expired_opps:
        pid        = opp.get("platform_id", "")
        human_vals = existing_rows.get(pid, {}).get("human_vals", [""] * len(HUMAN_COLUMNS))
        fill       = _fill(COLOR_EXPIRED)
        vals       = (["Expired", "", "EXPIRED", opp.get("organization", ""), opp.get("title", "")]
                      + [""] * (len(AGENT_COLUMNS) - 5) + human_vals)
        for col_idx, val in enumerate(vals, 1):
            c      = ws_archive.cell(row=arch_row, column=col_idx, value=val)
            c.fill = fill
            c.font = Font(size=9, color="888888")
        arch_row += 1

    # 9. Upload to Nextcloud
    url = upload_master(master_wb)

    # 10. Mirror to Google Drive (so Claude can read results each session)
    try:
        upload_to_gdrive(master_wb, scan_date)
    except Exception as e:
        log.warning(f"Google Drive mirror failed (non-critical): {e}")

    active_count = sum(1 for o in active_opps if o.get("_status") != "REJECTED")
    log.info(f"Combiner complete: {active_count} active bids in master tracker")
    return url


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
    from dotenv import load_dotenv
    load_dotenv()
    run_combiner(active_platforms={"Merx", "Biddingo", "Bonfire"})